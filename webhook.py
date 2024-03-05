from flask import Flask, request, abort, jsonify
import logging
import os
import json
from waitress import serve
from datetime import datetime
import calendar
import time
import xlrd
import openpyxl




logging.Formatter.converter = time.gmtime
log_path = os.environ.get('LOG_PATH', 'stdout')
log_level = os.environ.get('LOG_LEVEL', 20)

logging.basicConfig(filename=log_path, level=log_level, 
                    format='%(asctime)s %(levelname)s %(name)s %(message)s')

log = logging.getLogger(__name__) 

event = []
app = Flask(__name__)

## Method part

@app.route('/', methods=['GET'])
def home_page():
    if request.method == 'GET':
        print("Main page \n")
        return jsonify(event), 200
    else:
        abort(400)

@app.route('/webhook', methods=['POST'])
def webhook():
    if request.method == 'POST':
        wb = openpyxl.load_workbook('Book1.xlsx')
        # sheet = wb.get_sheet_by_name('Sheet1')
        payload = request.json

        # project_name = payload["project"]["namespace"]
        project_namespace = payload["project"]["namespace"]
        print(project_namespace)
        
        
        sheet_name = project_namespace

        if sheet_name in wb.sheetnames:
            print('Found sheet: ' + str(sheet_name) )
            sheet = wb[sheet_name]
            max_row = sheet.max_row
        else:
            print('Create new sheet for: ' + str(sheet_name) )
            wb.create_sheet(sheet_name)
            sheet = wb[sheet_name]
            sheet.append((["STT", "Commits", "Push event", "Merge event", "Date"]))
        
        print("Check if row is existed")
        date = datetime.today().strftime('%Y-%m-%d')
        found = False

        for row in sheet.rows:
            if row[4].value == date:
                found = True
                break

        print("Check type of request")
        if payload["object_kind"] == "merge_request":
            # datetime = payload["object_attributes"]["created_at"]
            # date = datetime.split(' ')[0]
            print("Merge")
            if found:
                print("Found today record")
                current_row = sheet.max_row
                pos = 'D' + str(current_row)
                print(sheet[pos].value)
                sheet[pos].value = int(sheet[pos].value) + 1
            else:
                sheet.append(([str(sheet.max_row - 1), 0, 0, 1, date]))
        elif payload["object_kind"] == "push":
            new_commits = len(payload["commits"])
            # date = datetime.today().strftime('%Y-%m-%d')
            print("push")
            if found:
                current_row = sheet.max_row
                push_pos = 'C' + str(current_row)
                commit_pos = 'B' + str(current_row)
                sheet[push_pos].value = int(sheet[push_pos].value) + 1
                sheet[commit_pos].value = int(sheet[commit_pos].value) + new_commits

            else:
                sheet.append(([str(sheet.max_row - 1), new_commits, 1, 0, date]))

        wb.save('./Book1.xlsx')
        return 'success', 200
    else:
        abort(400)

if __name__ == '__main__':
    print('------Start server------')
    serve(app, host="0.0.0.0", port=8080)